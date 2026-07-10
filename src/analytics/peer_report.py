import sqlite3
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATABASE = PROJECT_ROOT / "db" / "nifty100.db"
OUTPUT_FILE = PROJECT_ROOT / "output" / "peer_comparison.xlsx"

def load_peer_report_data():

    conn = sqlite3.connect(str(DATABASE))

    query = """
    SELECT
        pp.company_id,
        pp.peer_group_name,
        pp.metric,
        pp.value,
        pp.percentile_rank,
        pp.year,

        pg.is_benchmark,

        c.company_name

    FROM peer_percentiles pp

    LEFT JOIN peer_groups pg
        ON pp.company_id = pg.company_id

    LEFT JOIN companies c
        ON pp.company_id = c.id
    """

    df = pd.read_sql(query, conn)

    conn.close()

    return df

def export_peer_report(df):

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl"
    ) as writer:

        peer_groups = sorted(
            df["peer_group_name"]
            .dropna()
            .unique()
        )

        for group in peer_groups:

            group_df = df[
                df["peer_group_name"] == group
            ].copy()

            # Add median summary row
            summary = {}

            for column in group_df.columns:

                if pd.api.types.is_numeric_dtype(group_df[column]):
                    summary[column] = group_df[column].median()
                else:
                    summary[column] = ""

            summary["company_name"] = "Peer Group Median"

            group_df = pd.concat(
                [group_df, pd.DataFrame([summary])],
                ignore_index=True
            )

            group_df.to_excel(
                writer,
                sheet_name=group[:31],
                index=False
            )

    print("\nPeer comparison report exported successfully!")
    print(f"Location: {OUTPUT_FILE}")

def format_peer_report():

    wb = load_workbook(OUTPUT_FILE)

    green = PatternFill(
        start_color="90EE90",
        end_color="90EE90",
        fill_type="solid"
    )

    yellow = PatternFill(
        start_color="FFF59D",
        end_color="FFF59D",
        fill_type="solid"
    )

    red = PatternFill(
        start_color="FF9999",
        end_color="FF9999",
        fill_type="solid"
    )

    gold = PatternFill(
        start_color="FFD966",
        end_color="FFD966",
        fill_type="solid"
    )

    for sheet in wb.worksheets:

        headers = [cell.value for cell in sheet[1]]

        if "percentile_rank" not in headers:
            continue

        percentile_col = headers.index("percentile_rank") + 1

        benchmark_col = None

        if "is_benchmark" in headers:
            benchmark_col = headers.index("is_benchmark") + 1

        for row in range(2, sheet.max_row + 1):

            percentile = sheet.cell(row=row, column=percentile_col).value

            if isinstance(percentile, (int, float)):

                cell = sheet.cell(row=row, column=percentile_col)

                if percentile >= 75:
                    cell.fill = green

                elif percentile <= 25:
                    cell.fill = red

                else:
                    cell.fill = yellow

            if benchmark_col:

                if sheet.cell(row=row, column=benchmark_col).value == 1:

                    for col in range(1, sheet.max_column + 1):
                        sheet.cell(row=row, column=col).fill = gold

    wb.save(OUTPUT_FILE)

    print("Formatting applied successfully.")

if __name__ == "__main__":

    df = load_peer_report_data()

    export_peer_report(df)

    format_peer_report()